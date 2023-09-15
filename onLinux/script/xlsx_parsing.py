# -*- coding: utf-8 -*-

import sys
sys.path.append("./")
file_path = sys.argv[0]
import config

import os
import pymysql
import pandas as pd
from openpyxl import load_workbook
from sqlalchemy import create_engine, text
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils.cell import absolute_coordinate, coordinate_from_string
class xlsx_parsing:
    def __init__(self):
        self.pymysql_conn = pymysql.connect(host=config.DATABASE_CONFIG['host'],
                                            user=config.DATABASE_CONFIG['user'],
                                            password=config.DATABASE_CONFIG['password'],
                                            charset='utf8')
        self.pymysql_cur = self.pymysql_conn.cursor()

        self.pymysql_cur.execute(f"CREATE DATABASE if not exists {config.DATABASE_CONFIG['dbname']};")
        self.pymysql_conn.commit()

        self.engine = create_engine(f"mysql+pymysql://{config.DATABASE_CONFIG['user']}:{config.DATABASE_CONFIG['password']}@{config.DATABASE_CONFIG['host']}/{config.DATABASE_CONFIG['dbname']}")
        self.conn = self.engine.connect()

    def insert_data_sheet(self,xlsx_path):
        #현재 sample 액셀의 2번째 시트를 가져오기 위함
        input_xlsx_df = pd.read_excel(xlsx_path, sheet_name = 1)
        input_xlsx_df.columns = input_xlsx_df.iloc[8]
        input_xlsx_df = input_xlsx_df.iloc[9:,2:]
        
        input_xlsx_df.to_sql(name="data",con=self.engine, if_exists='replace', index=False)
        
        self.conn.execute(text("alter table data change `SUM(AMT)` `SUM(AMT)` DOUBLE;"))
        self.conn.commit()
    
    def create_result_sheet(self,xlsx_path):
        res = self.conn.execute(text("SELECT * FROM data;"))
        df = pd.DataFrame(res.fetchall())

        wb = load_workbook(xlsx_path, data_only=False)
        del wb['Data']

        ws = wb.create_sheet('Data')    
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)

        # 공백값 추가하여 원본과 같은 셀에서 시작
        ws.insert_rows(1,9)
        ws.insert_cols(1,2)

        # 함수식 추가
        ws["C7"] = "=COUNTA(C11:C1048576)"

        wb.save("./result.xlsx")

    def insert_result_sheet(self,xlsx_path,data_only):
        load_wb = load_workbook(xlsx_path, data_only=data_only)
        load_data_ws = load_wb['Data']
        load_ws = load_wb.active

        # 모든 행과 열 출력
        all_values = list()
        for row in load_ws.rows:
            row_value = list()
            for cell in row:
                row_value.append(cell.value)
            all_values.append(row_value)
        df = pd.DataFrame(all_values)

        # #empty dataframe
        # empty_df = pd.DataFrame([0,0],columns=['code'])
        
        # #code 열 추가
        # code4_index_df = df.iloc[27:161,8]
        # code6_index_df = df.iloc[27:161,9]
        # code4_index_df = code4_index_df.fillna(value=0)
        # code6_index_df = code6_index_df.fillna(value=0)
        # code46_index_df = code4_index_df + code6_index_df
        # code46_index_df.name='code'
        # code46_index_df = code46_index_df.to_frame()
        # code_index_df = pd.concat([empty_df,code46_index_df],axis=0)


        # 원수_column_code_df = df.iloc[25:27,11:60]
        # 원수_column_code_df.columns = config.원수_header_list
        # 출재_column_code_df = df.iloc[25:27,60:88]
        # 출재_column_code_df.columns = config.출재_header_list


        원수_df = df.iloc[28:161,11:60]
        원수_df.columns = config.원수_header_list
        # 원수_df = pd.concat([원수_column_code_df,원수_df],axis=0)
        # 원수_df = pd.concat([code_index_df,원수_df],axis=1)
        원수_df.to_sql(name=f"원수_tb_data_{str(data_only)}",con=self.engine, if_exists='replace', index=False)

        출재_df = df.iloc[28:161,60:88]
        출재_df.columns = config.출재_header_list
        # 출재_df = pd.concat([출재_column_code_df,출재_df],axis=0)
        # 출재_df = pd.concat([code_index_df,출재_df],axis=1)
        출재_df.to_sql(name=f"출재_tb_data_{str(data_only)}",con=self.engine, if_exists='replace', index=False)
    
    def insert_result_sheet_coordinate(self,xlsx_path,data_only):
        load_wb = load_workbook(xlsx_path, data_only=data_only)
        load_data_ws = load_wb['Data']
        load_ws = load_wb.active

        # 모든 행과 열 출력
        all_values = list()
        for row in load_ws.rows:
            for cell in row:
                row_dict = dict()
                row_dict['coordinate'] = cell.coordinate
                row_dict['absolute_coordinate'] = absolute_coordinate(cell.coordinate)
                row_dict['column'] = coordinate_from_string(cell.coordinate)[0]
                row_dict['row'] = coordinate_from_string(cell.coordinate)[1]
                row_dict['value'] = cell.value
                all_values.append(row_dict)
        df = pd.DataFrame(all_values)
        df.to_sql(name=f"coordinate_tb_data_{str(data_only)}",con=self.engine, if_exists='replace', index=False)
    
    def db_close(self):
        self.pymysql_conn.close()
        self.conn.close()

    def main(self,xlsx_path):
        print(">> insert_data_sheet")
        self.insert_data_sheet(xlsx_path)
        print(">> create_result_sheet")
        self.create_result_sheet(xlsx_path)
        print(">> insert_result_sheet")
        self.insert_result_sheet(xlsx_path,True)
        self.insert_result_sheet(xlsx_path,False)
        print(">> insert_result_sheet_coordinate")
        self.insert_result_sheet_coordinate(xlsx_path,True)
        self.insert_result_sheet_coordinate(xlsx_path,False)
        self.db_close()
        print("complete")
