# -*- coding: utf-8 -*-

import sys
sys.path.append("./")
file_path = sys.argv[0]
import config

import os
import pymysql
import pandas as pd
from openpyxl import load_workbook
from sqlalchemy import create_engine, text
from openpyxl.utils.dataframe import dataframe_to_rows

def upload_xlsx(input_xlsx_path):
        input_xlsx_df = pd.read_excel(input_xlsx_path,sheet_name = 1)
        input_xlsx_df.columns = input_xlsx_df.iloc[8]
        input_xlsx_df = input_xlsx_df.iloc[9:,2:]
        
        input_xlsx_df.to_sql(name="data",con=engine, if_exists='replace', index=False)
        
        conn.execute(text("alter table data change `SUM(AMT)` `SUM(AMT)` DOUBLE;"))
        conn.commit()

if __name__ == "__main__":
    if len(sys.argv) != 2:
        print("엑셀 파일을 넣으세요 ex) python .\main.py .\A.xlsx")
        sys.exit()
    else:
        pymysql_conn = pymysql.connect(host=config.DATABASE_CONFIG['host'],
                       user=config.DATABASE_CONFIG['user'],
                       password=config.DATABASE_CONFIG['password'],
                       charset='utf8')

        pymysql_cur = pymysql_conn.cursor()
        pymysql_cur.execute(f"CREATE DATABASE if not exists {config.DATABASE_CONFIG['dbname']};")
        pymysql_conn.commit()

        engine = create_engine(f"mysql+pymysql://{config.DATABASE_CONFIG['user']}:{config.DATABASE_CONFIG['password']}@{config.DATABASE_CONFIG['host']}/{config.DATABASE_CONFIG['dbname']}")
        conn = engine.connect()

        print("DB 생성 혹은 유지 완료")
        #################################################################

        input_xlsx_path = sys.argv[1]
        upload_xlsx(input_xlsx_path)
        print("DB 업로드 완료")
        #################################################################

        res = conn.execute(text("SELECT * FROM data;"))
        df = pd.DataFrame(res.fetchall())

        wb = load_workbook(input_xlsx_path, data_only=False)
        del wb['Data']

        ws = wb.create_sheet('Data')    
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)

        # 공백값 추가하여 원본과 같은 셀에서 시작
        ws.insert_rows(1,9)
        ws.insert_cols(1,2)

        # 함수식 추가
        ws["C7"] = "=COUNTA(C11:C1048576)"

        wb.save("./result.xlsx")
        print("엑셀 시트 생성 완료")
        #################################################################
        
        #data_only = True 로 된 엑셀 시트 로드
        load_wb = load_workbook(input_xlsx_path, data_only=True)
        load_data_ws = load_wb['Data']
        load_ws = load_wb.active

        # 모든 행과 열 출력
        all_values = list()
        for row in load_ws.rows:
            row_value = list()
            for cell in row:
                row_value.append(cell.value)
            all_values.append(row_value)
        df = pd.DataFrame(all_values)

        header_list = list(df.iloc[24,11:21])
        new_df = df.iloc[28:161,11:21]
        new_df.columns = header_list
        new_df.to_sql(name="result",con=engine, if_exists='replace', index=False)
        print("result 시트 DB 업로드 완료")