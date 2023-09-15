# -*- coding: utf-8 -*-

import sys
sys.path.append("./")
import config

import os
import pandas as pd
from openpyxl import load_workbook

#sqlalchemy 연동
from sqlalchemy import create_engine
db_connection_str = f"mysql+pymysql://{config.DATABASE_CONFIG['user']}:{config.DATABASE_CONFIG['password']}@{config.DATABASE_CONFIG['host']}/{config.DATABASE_CONFIG['dbname']}"
db_connection = create_engine(db_connection_str)


def excel_to_sql(file_name):
    #openpyxl로 액셀 다루기
    wb = load_workbook(f"./output_file/{file_name}", data_only=False)
    ws_names = wb.sheetnames
    for ws_name in ws_names:
        ws = wb[ws_name]

        # 모든 행과 열 출력
        sheet_list = list()
        for row in ws.rows:
            row_list = list()
            for cell in row:
                row_list.append(cell.value)
            sheet_list.append(row_list)
        sheet_df = pd.DataFrame(sheet_list)
        # sheet_df = sheet_df.dropna(axis=0, how='all')
        # sheet_df.columns=sheet_df.iloc[0]

        # sqlalchemy 활용하여 mysql 적재
        sheet_df.to_sql(name=file_name.rstrip(".xlsm")+"_"+ws_name,con=db_connection, if_exists='append', index=False)